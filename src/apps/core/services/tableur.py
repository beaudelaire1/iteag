"""Lecture et écriture de tableurs — CSV et Excel, sans connaître aucun domaine.

Ce module ne sait rien des étudiants ni des livres : il transforme des lignes
en dictionnaires et des dictionnaires en fichier. Les colonnes, les règles et
les validations vivent auprès des modèles qu'elles décrivent.

Trois pièges de terrain sont traités ici, parce qu'ils ne se voient qu'à
l'usage et qu'aucun d'eux ne lève d'erreur :

1. **Le séparateur.** Excel en France écrit des CSV séparés par des
   points-virgules. Lu avec des virgules, le fichier donne une seule colonne
   par ligne — l'import « réussit » et n'importe rien.
2. **Le BOM.** Excel préfixe ses CSV UTF-8 d'une marque d'ordre d'octets. Sans
   « utf-8-sig », la première en-tête devient « ﻿nom » et la colonne
   paraît manquante alors qu'elle est là.
3. **Les nombres devenus texte.** Un numéro étudiant « ETU2026001 » reste une
   chaîne, mais un code purement numérique est relu par Excel comme un flottant
   et ressort « 2026001.0 ». Les valeurs sont donc normalisées à la lecture.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Callable, Iterable, Sequence
from dataclasses import dataclass, field
from typing import Any

from django.http import HttpResponse

SEPARATEURS = [";", ",", "\t"]
EXTENSIONS_TABLEUR = (".xlsx", ".xlsm")


class FichierIllisible(Exception):
    """Le fichier n'est ni un CSV ni un classeur exploitable."""


@dataclass(frozen=True)
class Colonne:
    """Une colonne du gabarit : son en-tête, ce qu'elle attend, un exemple."""

    nom: str
    aide: str = ""
    requise: bool = False
    exemple: str = ""


@dataclass
class Rapport:
    """Ce qu'un import a fait, ou aurait fait.

    Les erreurs portent le numéro de ligne du tableur — celui que l'utilisateur
    voit dans Excel, en-tête comprise — sans quoi « ligne 12 » ne désigne rien
    de repérable à l'écran.
    """

    crees: int = 0
    mis_a_jour: int = 0
    ignores: int = 0
    erreurs: list[tuple[int, str]] = field(default_factory=list)

    @property
    def total_traite(self) -> int:
        return self.crees + self.mis_a_jour

    @property
    def est_en_echec(self) -> bool:
        return bool(self.erreurs)

    def resume(self) -> str:
        if self.est_en_echec:
            return f"{len(self.erreurs)} ligne(s) en erreur : rien n'a été enregistré."
        morceaux = []
        if self.crees:
            morceaux.append(f"{self.crees} création(s)")
        if self.mis_a_jour:
            morceaux.append(f"{self.mis_a_jour} mise(s) à jour")
        if self.ignores:
            morceaux.append(f"{self.ignores} ligne(s) vide(s) ignorée(s)")
        return " · ".join(morceaux) if morceaux else "Aucune ligne exploitable."


def _texte(valeur: Any) -> str:
    """Normalise une cellule en chaîne utilisable.

    Excel relit un code numérique en flottant : « 2026001 » revient
    « 2026001.0 », et la comparaison avec la base échoue sans rien dire.
    """
    if valeur is None:
        return ""
    if isinstance(valeur, float) and valeur.is_integer():
        return str(int(valeur))
    return str(valeur).strip()


def _lignes_du_csv(contenu: bytes) -> list[list[str]]:
    # « utf-8-sig » retire la marque d'ordre d'octets qu'Excel ajoute ; le repli
    # cp1252 couvre les fichiers enregistrés en « CSV Windows ».
    for encodage in ("utf-8-sig", "cp1252"):
        try:
            texte = contenu.decode(encodage)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise FichierIllisible("Encodage du fichier non reconnu : enregistrez-le en UTF-8.")

    premiere_ligne = texte.splitlines()[0] if texte.splitlines() else ""
    separateur = max(SEPARATEURS, key=premiere_ligne.count)
    if premiere_ligne.count(separateur) == 0:
        separateur = ";"
    return [ligne for ligne in csv.reader(io.StringIO(texte), delimiter=separateur)]


def _lignes_du_classeur(contenu: bytes) -> list[list[Any]]:
    from openpyxl import load_workbook

    try:
        classeur = load_workbook(io.BytesIO(contenu), read_only=True, data_only=True)
    except Exception as erreur:  # noqa: BLE001 — openpyxl lève des types variés
        raise FichierIllisible("Classeur illisible : réenregistrez-le au format .xlsx.") from erreur
    feuille = classeur[classeur.sheetnames[0]]
    return [list(ligne) for ligne in feuille.iter_rows(values_only=True)]


def lire(fichier) -> list[dict[str, str]]:
    """Retourne une liste de dictionnaires en-tête → valeur.

    Les en-têtes sont normalisées en minuscules sans espaces de bord : un
    « Nom » saisi « nom  » désigne la même colonne, et l'utilisateur n'a pas à
    deviner la casse attendue.
    """
    contenu = fichier.read()
    nom = getattr(fichier, "name", "") or ""

    lignes = _lignes_du_classeur(contenu) if nom.lower().endswith(EXTENSIONS_TABLEUR) else _lignes_du_csv(contenu)
    if not lignes:
        raise FichierIllisible("Le fichier est vide.")

    entetes = [_texte(cellule).lower() for cellule in lignes[0]]
    if not any(entetes):
        raise FichierIllisible("La première ligne doit porter les en-têtes de colonnes.")

    resultat = []
    for brute in lignes[1:]:
        valeurs = [_texte(cellule) for cellule in brute]
        # Les lignes plus courtes que l'en-tête sont complétées, les plus
        # longues tronquées : « strict » interdirait un fichier dont la
        # dernière colonne est simplement vide — cas le plus courant d'Excel.
        valeurs = (valeurs + [""] * len(entetes))[: len(entetes)]
        resultat.append({e: v for e, v in zip(entetes, valeurs, strict=True) if e})
    return resultat


def _reponse(nom_fichier: str, type_mime: str, contenu: bytes) -> HttpResponse:
    reponse = HttpResponse(contenu, content_type=type_mime)
    reponse["Content-Disposition"] = f'attachment; filename="{nom_fichier}"'
    return reponse


def _cellule_sure(valeur: Any) -> str:
    """Neutralise ce qu'un tableur interpréterait comme une formule.

    Une valeur commençant par « = » ou « + » est exécutée à l'ouverture : un
    nom d'étudiant malveillant deviendrait une commande. Le préfixe apostrophe
    est la parade retenue par le reste du projet pour les exports CSV.
    """
    texte = _texte(valeur)
    return f"'{texte}" if texte.lstrip().startswith(("=", "+", "-", "@")) else texte


def ecrire_csv(nom_fichier: str, colonnes: Sequence[str], lignes: Iterable[Sequence[Any]]) -> HttpResponse:
    tampon = io.StringIO()
    # Point-virgule : c'est ce qu'Excel attend d'un CSV en français, et un
    # export qu'il faut réparer à la main n'est pas un export.
    graveur = csv.writer(tampon, delimiter=";")
    graveur.writerow(colonnes)
    for ligne in lignes:
        graveur.writerow([_cellule_sure(valeur) for valeur in ligne])
    return _reponse(nom_fichier, "text/csv; charset=utf-8-sig", tampon.getvalue().encode("utf-8-sig"))


def ecrire_xlsx(nom_fichier: str, colonnes: Sequence[str], lignes: Iterable[Sequence[Any]]) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Données"
    feuille.append(list(colonnes))
    for cellule in feuille[1]:
        cellule.font = Font(bold=True)
    for ligne in lignes:
        feuille.append([_cellule_sure(valeur) for valeur in ligne])

    # Une largeur approchée vaut mieux qu'une colonne de dièses à l'ouverture.
    for index, entete in enumerate(colonnes, start=1):
        feuille.column_dimensions[feuille.cell(row=1, column=index).column_letter].width = max(14, len(entete) + 4)
    feuille.freeze_panes = "A2"

    tampon = io.BytesIO()
    classeur.save(tampon)
    return _reponse(
        nom_fichier,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        tampon.getvalue(),
    )


def gabarit(nom_fichier: str, colonnes: Sequence[Colonne], format_demande: str = "xlsx") -> HttpResponse:
    """Un gabarit vide, en-têtes et ligne d'exemple.

    L'exemple est indispensable : sans lui, « date_naissance » se remplit dans
    cinq formats différents selon la personne, et l'import rejette des lignes
    que rien ne distinguait à l'œil.
    """
    entetes = [colonne.nom for colonne in colonnes]
    exemple = [colonne.exemple for colonne in colonnes]
    lignes = [exemple] if any(exemple) else []
    if format_demande == "csv":
        return ecrire_csv(nom_fichier, entetes, lignes)
    return ecrire_xlsx(nom_fichier, entetes, lignes)


def valider_entetes(donnees: list[dict[str, str]], colonnes: Sequence[Colonne]) -> list[str]:
    """Colonnes obligatoires absentes du fichier déposé."""
    presentes = set(donnees[0].keys()) if donnees else set()
    return [colonne.nom for colonne in colonnes if colonne.requise and colonne.nom not in presentes]


def ligne_vide(ligne: dict[str, str]) -> bool:
    """Une ligne où tout est blanc : Excel en produit des milliers en fin de feuille."""
    return not any(valeur.strip() for valeur in ligne.values())


@dataclass(frozen=True)
class Schema:
    """Ce qu'une entité expose à l'import et à l'export.

    « importer_ligne » reçoit un dictionnaire et retourne « cree » (booléen).
    Toute erreur métier se signale par une ValidationError : le rapport la
    rattache à son numéro de ligne, et la transaction complète est annulée.
    """

    cle: str
    libelle: str
    colonnes: Sequence[Colonne]
    exporter: Callable[[], Iterable[Sequence[Any]]]
    importer_ligne: Callable[[dict[str, str]], bool] | None = None

    @property
    def entetes(self) -> list[str]:
        return [colonne.nom for colonne in self.colonnes]

    @property
    def importable(self) -> bool:
        return self.importer_ligne is not None
